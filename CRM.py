import streamlit as st
import pandas as pd
import altair as alt
from openpyxl import load_workbook

# --- Path Excel ---
file_path = r"C:\Users\user\Documents\CRM UPGRADE TEXTEK\CRM Analyst.xlsx"

# --- Helper Aman: Simpan ke Sheet tanpa Overwrite Workbook ---
def save_to_excel_safely(df, file_path, sheet_name):
    book = load_workbook(file_path)
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        writer.book = book
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()

# --- Pilih Sheet ---
sheet_options = [
    "VIP BUYER",
    "Kategori Buyer",
    "Marketing Ads",
    "Pertumbuhan Pelanggan",
    "Produk Populer",
    "Produk Favorit Customer"
]
current_sheet = st.selectbox("Pilih Sheet CRM", sheet_options)

# =========================
# SHEET 1: VIP BUYER
# =========================
if current_sheet == "VIP BUYER":
    df = pd.read_excel(file_path, sheet_name="VIP BUYER", engine='openpyxl')
    df.columns = df.columns.map(lambda x: str(x).strip() if x is not None else x)

    # Bersihkan Total Transaksi
    df['Total Transaksi'] = df['Total Transaksi'].replace(['-', '', ' '], 0)
    df['Total Transaksi'] = df['Total Transaksi'].replace('[Rp$,]', '', regex=True)
    df['Total Transaksi'] = pd.to_numeric(df['Total Transaksi'], errors='coerce').fillna(0)

    # Session state
    if 'vip_data' not in st.session_state:
        st.session_state.vip_data = df.copy()

    # --- Form Tambah Data ---
    st.subheader("Tambah Data Baru")
    with st.form("add_form"):
        nama = st.text_input("Nama Pelanggan")
        jumlah = st.number_input("Jumlah Transaksi", min_value=0, step=1)
        total = st.number_input("Total Transaksi (Rp)", min_value=0, step=1000)
        submitted = st.form_submit_button("Tambah")
        if submitted:
            if nama.strip() != "":
                new_row = pd.DataFrame({
                    'Nama Pelanggan': [nama],
                    'Jumlah Transaksi': [jumlah],
                    'Total Transaksi': [total]
                })
                st.session_state.vip_data = pd.concat([st.session_state.vip_data, new_row], ignore_index=True)
                st.success(f"✅ {nama} berhasil ditambahkan!")
            else:
                st.warning("Nama pelanggan tidak boleh kosong.")

    # --- Form Hapus Data ---
    st.subheader("Hapus Data")
    if len(st.session_state.vip_data) > 0:
        delete_name = st.selectbox("Pilih Nama Pelanggan yang akan dihapus", st.session_state.vip_data['Nama Pelanggan'])
        if st.button("Hapus"):
            st.session_state.vip_data = st.session_state.vip_data[st.session_state.vip_data['Nama Pelanggan'] != delete_name]
            st.success(f"🗑️ {delete_name} berhasil dihapus!")

    # --- Tampilkan Tabel VIP BUYER ---
    df_display = st.session_state.vip_data.sort_values(by='Total Transaksi', ascending=False).reset_index(drop=True)
    df_display.index = range(1, len(df_display)+1)

    # Highlight Top 10
    top10_idx = df_display.index[:10]
    def highlight_top10_all(row):
        if row.name in top10_idx[:3]:
            return ['background-color: lightgreen'] * len(row)
        elif row.name in top10_idx[3:10]:
            return ['background-color: lightyellow'] * len(row)
        else:
            return [''] * len(row)

    st.subheader("Tabel VIP BUYER")
    st.dataframe(df_display.style.format({'Total Transaksi': 'Rp {:,.0f}'}).apply(highlight_top10_all, axis=1))


# =========================
# SHEET 2: KATEGORI BUYER
# =========================
elif current_sheet == "Kategori Buyer":
    df = pd.read_excel(file_path, sheet_name="Kategori Buyer", engine='openpyxl')
    df.columns = df.columns.str.strip()
    df = df[['Nama Customer','Repeat Status','Buyer Status']]
    df.index = range(1, len(df)+1)

    # Highlight Buyer Status
    buyer_color = ['#FFC0CB','#ADD8E6','#90EE90','#FFFFE0','#FFA07A','#D8BFD8','#F0E68C']
    buyer_categories = ['Active Buyer','Cooling Buyer','Dormant Buyer','Inactive Buyer','Lost Buyer','Very Inactive Buyer','Warm Buyer']
    buyer_map = {v: buyer_color[i] for i,v in enumerate(buyer_categories)}
    def highlight_buyer(x):
        return [f'background-color: {buyer_map.get(v,"")}' if col=='Buyer Status' else '' for col,v in zip(x.index,x)]

    st.subheader("Tabel Kategori Buyer")
    st.dataframe(df.style.apply(highlight_buyer, axis=1))

    # Ringkasan Kategori Buyer 
    st.subheader("Ringkasan Jumlah per Kategori")
    jumlah_per_kategori = {
        "Active Buyer":5,
        "Cooling Buyer":8,
        "Dormant Buyer":5,
        "Inactive Buyer":6,
        "Lost Buyer":22,
        "Very Inactive Buyer":5,
        "Warm Buyer":12
    }
    for k,v in jumlah_per_kategori.items():
        st.write(f"{k} → {v}")

    st.subheader("Rata-rata Jarak Pembelian per Kategori")
    rata_per_kategori = {
        "Active Buyer":22,
        "Cooling Buyer":8,
        "Warm Buyer":14,
        "Lost Buyer":6,
        "Dormant Buyer":35,
        "Very Inactive Buyer":10,
        "Inactive Buyer":18
    }
    for k,v in rata_per_kategori.items():
        st.write(f"{k} → {v} hari")

    # =========================
    # Ringkasan Buyer Status - Versi Formal
    # =========================
    st.subheader("Ringkasan Buyer Status")
    st.markdown("""
**Repeat Buyer:** Pelanggan yang telah melakukan transaksi lebih dari satu kali, menunjukkan loyalitas dan potensi tinggi untuk pembelian berulang.  
**Non-Repeat Buyer:** Pelanggan yang baru melakukan satu kali transaksi, memerlukan perhatian untuk meningkatkan kemungkinan repeat order.

---

🟢 **Active Buyer (≤15 hari)**  
Pelanggan yang baru saja melakukan transaksi atau masih sangat aktif.  
➡️ Potensi tinggi untuk pembelian berulang; cukup pertahankan kualitas layanan dan komunikasi.

🟡 **Warm Buyer (16–30 hari)**  
Pelanggan yang masih menunjukkan minat, namun frekuensi pembelian menurun.  
➡️ Strategi ringan seperti promo atau follow-up ramah dapat meningkatkan peluang transaksi berikutnya.

🟠 **Cooling Buyer (31–45 hari)**  
Frekuensi pembelian mulai menurun secara signifikan.  
➡️ Dibutuhkan stimulasi seperti diskon terbatas, pengingat produk baru, atau kampanye “kami merindukan Anda”.

🔵 **Dormant Buyer (46–60 hari)**  
Pelanggan yang sudah cukup lama tidak melakukan pembelian.  
➡️ Strategi reaktivasi diperlukan melalui komunikasi personal, voucher, atau pesan langsung.

🟣 **Inactive Buyer (61–90 hari)**  
Pelanggan dalam fase tidak aktif yang lebih lama.  
➡️ Dibutuhkan pendekatan berbasis nilai atau emosional, bukan sekadar promosi.

⚫ **Very Inactive Buyer (91–120 hari)**  
Pelanggan hampir tidak aktif, namun masih memungkinkan untuk dikembalikan.  
➡️ Perlu alasan kuat untuk melakukan pembelian kembali, misalnya produk baru atau penawaran khusus.

🔴 **Lost Buyer (>120 hari)**  
Pelanggan yang kemungkinan besar sudah hilang.  
➡️ Data ini dapat digunakan untuk analisis churn dan strategi pencegahan agar pelanggan lain tidak menjadi lost.
""")

# =========================
# SHEET 3: MARKETING ADS
# =========================
elif current_sheet == "Marketing Ads":
    df_marketing = pd.read_excel(file_path, sheet_name="Marketing Ads", engine='openpyxl')
    df_marketing.columns = df_marketing.columns.str.strip()
    df_marketing['Jumlah'] = pd.to_numeric(df_marketing['Jumlah'], errors='coerce').fillna(0)

    st.subheader("Distribusi Channel Marketing")
    st.write("Berikut adalah distribusi jumlah customer berdasarkan channel marketing:")

    # --- Warna konsisten per channel ---
    color_map = {
        "Instagram": "#1F77B4",   # biru muda
        "Facebook": "#2E86AB",    # biru tua
        "WhatsApp Ads": "#28B463",# hijau
        "Non-Ads": "#A9A9A9",     # abu
        "TikTok": "#E74C3C",      # merah
        "Website": "#F1C40F",     # kuning
    }

    colors = [color_map.get(ch, "#CCCCCC") for ch in df_marketing['Channel']]

    # --- Pie Chart dengan label & legend ---
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots()
    wedges, texts, autotexts = ax.pie(
        df_marketing['Jumlah'],
        labels=None,  # label diganti legend biar clean
        autopct='%1.1f%%',
        startangle=90,
        colors=colors,
        textprops={'fontsize': 10, 'color': 'white'}
    )

    for t in autotexts:
        t.set_color('black')
        t.set_fontweight('bold')

    ax.axis('equal')
    ax.legend(
        wedges,
        df_marketing['Channel'],
        title="Channel Marketing",
        loc="center left",
        bbox_to_anchor=(1, 0, 0.5, 1)
    )

    st.pyplot(fig)

    # --- Ringkasan Insight ---
    top_channel = df_marketing.loc[df_marketing['Jumlah'].idxmax(), 'Channel']
    st.subheader("Ringkasan Insight")
    st.write(f"Dari data marketing ads, **{top_channel}** adalah yang paling efektif menarik customer karena memiliki persentase tertinggi.")

# =========================
# SHEET 4: PERTUMBUHAN PELANGGAN
# =========================
elif current_sheet == "Pertumbuhan Pelanggan":
    df = pd.read_excel(file_path, sheet_name="Pertumbuhan Pelanggan", engine='openpyxl')
    df.columns = df.columns.str.strip()
    df['Jumlah Pelanggan'] = pd.to_numeric(df['Jumlah Pelanggan'], errors='coerce').fillna(0)

    # Hapus baris kosong / None
    df = df.dropna(subset=['Bulan','Jumlah Pelanggan']).reset_index(drop=True)

    # Session state
    if 'growth_data' not in st.session_state:
        st.session_state.growth_data = df.copy()

    # --- Form Tambah Data ---
    st.subheader("Tambah Data Pertumbuhan")
    with st.form("add_growth"):
        bulan = st.text_input("Bulan")
        jumlah = st.number_input("Jumlah Pelanggan", min_value=0, step=1)
        submitted = st.form_submit_button("Tambah")
        if submitted:
            new_row = pd.DataFrame({'Bulan':[bulan],'Jumlah Pelanggan':[jumlah]})
            st.session_state.growth_data = pd.concat([st.session_state.growth_data, new_row], ignore_index=True)
            st.success(f"{bulan} berhasil ditambahkan!")

    # --- Form Hapus Data ---
    st.subheader("Hapus Data Pertumbuhan")
    if len(st.session_state.growth_data) > 0:
        delete_bulan = st.selectbox("Pilih Bulan yang akan dihapus", st.session_state.growth_data['Bulan'])
        if st.button("Hapus Bulan"):
            st.session_state.growth_data = st.session_state.growth_data[st.session_state.growth_data['Bulan'] != delete_bulan]
            st.success(f"{delete_bulan} berhasil dihapus!")

    # --- Tampilkan Tabel ---
    df_display = st.session_state.growth_data.sort_values(by='Jumlah Pelanggan', ascending=False).reset_index(drop=True)
    df_display.index = range(len(df_display))  # pakai 0-based supaya logika trend bener

    # Trend: 🔼 top 3 hijau, nomor 4 strip kuning, sisanya 🔽 merah
    df_display['Trend'] = ''
    for i in range(len(df_display)):
        if i < 3:
            df_display.loc[i, 'Trend'] = '🔼'
        elif i == 3:
            df_display.loc[i, 'Trend'] = '—'
        else:
            df_display.loc[i, 'Trend'] = '🔽'

    # Highlight berdasarkan Trend
    def highlight_trend(row):
        if row['Trend'] == '🔼':
            return ['background-color: lightgreen']*len(row)
        elif row['Trend'] == '—':
            return ['background-color: lightyellow']*len(row)
        elif row['Trend'] == '🔽':
            return ['background-color: lightcoral']*len(row)
        else:
            return ['']*len(row)

    st.subheader("Tabel Pertumbuhan Pelanggan")
    st.dataframe(df_display.style.apply(highlight_trend, axis=1))

# =========================
# SHEET 5L: PRODUK POPULER
# =========================
elif current_sheet == "Produk Populer":
    import io

    # --- Load data ---
    df = pd.read_excel(file_path, sheet_name="Produk Populer", engine='openpyxl')
    df.columns = df.columns.str.strip()
    df['Jumlah Pembelian'] = pd.to_numeric(df['Jumlah Pembelian'], errors='coerce').fillna(0)
    df_display = df.sort_values(by='Jumlah Pembelian', ascending=False).reset_index(drop=True)
    df_display.index = range(1, len(df_display)+1)

    # --- Highlight Top 5 ---
    top5_idx = df_display.index[:5]
    def highlight_top5(row):
        if row.name in top5_idx:
            return ['background-color: lightgreen'] * len(row)
        else:
            return [''] * len(row)

    st.subheader("Tabel Produk Populer")
    st.dataframe(df_display.style.apply(highlight_top5, axis=1))

    # ======================
    # CRUD SECTION
    # ======================
    st.markdown("---")
    st.subheader("🧩 Edit Data Produk Populer")

    crud_action = st.radio("Pilih Aksi", ["Tambah Data", "Edit Data", "Hapus Data"], horizontal=True)

    # --- TAMBAH DATA ---
    if crud_action == "Tambah Data":
        st.markdown("### ➕ Tambah Produk Baru")
        new_name = st.text_input("Jenis Produk")
        new_value = st.number_input("Jumlah Pembelian", min_value=0, step=1)
        if st.button("Tambah"):
            if new_name:
                new_row = pd.DataFrame({"Jenis Produk": [new_name], "Jumlah Pembelian": [new_value]})
                df = pd.concat([df, new_row], ignore_index=True)
                df = df.sort_values(by="Jumlah Pembelian", ascending=False)
                df.to_excel(file_path, sheet_name="Produk Populer", index=False, engine='openpyxl')
                st.success(f"✅ Produk '{new_name}' berhasil ditambahkan!")
                st.rerun()
            else:
                st.warning("Isi nama produk terlebih dahulu.")

    # --- EDIT DATA ---
    elif crud_action == "Edit Data":
        st.markdown("### ✏️ Edit Produk")
        product_list = df['Jenis Produk'].tolist()
        selected_product = st.selectbox("Pilih Produk untuk Diedit", product_list)
        selected_row = df[df['Jenis Produk'] == selected_product].iloc[0]
        new_name = st.text_input("Jenis Produk", value=selected_row['Jenis Produk'])
        new_value = st.number_input("Jumlah Pembelian", value=int(selected_row['Jumlah Pembelian']), min_value=0, step=1)
        if st.button("Simpan Perubahan"):
            df.loc[df['Jenis Produk'] == selected_product, 'Jenis Produk'] = new_name
            df.loc[df['Jenis Produk'] == selected_product, 'Jumlah Pembelian'] = new_value
            df = df.sort_values(by="Jumlah Pembelian", ascending=False)
            df.to_excel(file_path, sheet_name="Produk Populer", index=False, engine='openpyxl')
            st.success(f"✅ Produk '{selected_product}' berhasil diperbarui!")
            st.rerun()

    # --- HAPUS DATA ---
    elif crud_action == "Hapus Data":
        st.markdown("### 🗑️ Hapus Produk")
        product_list = df['Jenis Produk'].tolist()
        selected_product = st.selectbox("Pilih Produk untuk Dihapus", product_list)
        if st.button("Hapus"):
            df = df[df['Jenis Produk'] != selected_product]
            df = df.sort_values(by="Jumlah Pembelian", ascending=False)
            df.to_excel(file_path, sheet_name="Produk Populer", index=False, engine='openpyxl')
            st.success(f"🗑️ Produk '{selected_product}' berhasil dihapus!")
            st.rerun()

# =========================
# SHEET 6: PRODUK FAVORIT CUSTOMER
# =========================
elif current_sheet == "Produk Favorit Customer":
    df = pd.read_excel(file_path, sheet_name="Produk Favorit Customer", engine='openpyxl')
    df.columns = df.columns.str.strip()
    df['Jumlah Dibeli'] = pd.to_numeric(df['Jumlah Dibeli'], errors='coerce').fillna(0).astype(int)

    if 'fav_data' not in st.session_state:
        st.session_state.fav_data = df

    # --- tampilkan tabel dengan nomor mulai dari 1
    df_display = st.session_state.fav_data.copy()
    df_display.index = range(1, len(df_display) + 1)

    st.subheader("Tabel Produk Favorit Customer")
    st.dataframe(df_display, use_container_width=True)

    # --- Tabs CRUD (biar gak numpuk)
    tab_tambah, tab_edit, tab_hapus = st.tabs(["➕ Tambah", "✏️ Edit", "🗑️ Hapus"])

    # --- TAMBAH DATA ---
    with tab_tambah:
        with st.form("form_tambah_fav", clear_on_submit=True):
            nama_customer = st.text_input("Nama Customer")
            produk_fav = st.text_input("Produk Favorit")
            jumlah_dibeli = st.number_input("Jumlah Dibeli", min_value=0, step=1)
            submitted = st.form_submit_button("Tambah")

            if submitted:
                new_row = pd.DataFrame({
                    "Nama Customer": [nama_customer],
                    "Produk Favorit": [produk_fav],
                    "Jumlah Dibeli": [jumlah_dibeli]
                })
                st.session_state.fav_data = pd.concat([st.session_state.fav_data, new_row], ignore_index=True)
                st.success("✅ Data berhasil ditambahkan!")
                st.rerun()

    # --- EDIT DATA ---
    with tab_edit:
        if not st.session_state.fav_data.empty:
            row_to_edit = st.selectbox(
                "Pilih baris untuk diedit",
                options=st.session_state.fav_data.index,
                format_func=lambda i: f"{st.session_state.fav_data.at[i, 'Nama Customer']} - {st.session_state.fav_data.at[i, 'Produk Favorit']}"
            )

            nama_customer_edit = st.text_input("Nama Customer", st.session_state.fav_data.at[row_to_edit, 'Nama Customer'])
            produk_fav_edit = st.text_input("Produk Favorit", st.session_state.fav_data.at[row_to_edit, 'Produk Favorit'])
            jumlah_dibeli_edit = st.number_input("Jumlah Dibeli", min_value=0, step=1, value=int(st.session_state.fav_data.at[row_to_edit, 'Jumlah Dibeli']))
            
            if st.button("Simpan Perubahan"):
                st.session_state.fav_data.at[row_to_edit, 'Nama Customer'] = nama_customer_edit
                st.session_state.fav_data.at[row_to_edit, 'Produk Favorit'] = produk_fav_edit
                st.session_state.fav_data.at[row_to_edit, 'Jumlah Dibeli'] = jumlah_dibeli_edit
                st.success("✅ Data berhasil diperbarui!")
                st.rerun()

    # --- HAPUS DATA ---
    with tab_hapus:
        if not st.session_state.fav_data.empty:
            row_to_delete = st.selectbox(
                "Pilih baris untuk dihapus",
                options=st.session_state.fav_data.index,
                format_func=lambda i: f"{st.session_state.fav_data.at[i, 'Nama Customer']} - {st.session_state.fav_data.at[i, 'Produk Favorit']}"
            )
            if st.button("Hapus Data"):
                st.session_state.fav_data = st.session_state.fav_data.drop(row_to_delete).reset_index(drop=True)
                st.success("🗑️ Data berhasil dihapus!")
                st.rerun()

    # Simpan ke Excel (overwrite hanya sheet ini)
    from openpyxl import load_workbook
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        st.session_state.fav_data.to_excel(writer, sheet_name="Produk Favorit Customer", index=False)
